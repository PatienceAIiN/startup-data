import csv
import io
import os
import tempfile
from datetime import datetime
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import structlog
from app.services.r2_service import upload_file_to_r2, generate_presigned_url

log = structlog.get_logger()

EXPORT_COLUMNS = [
    ("CIN", "cin"),
    ("Company Name", "company_name"),
    ("Status", "company_status"),
    ("ROC Code", "roc_code"),
    ("Category", "company_category"),
    ("Date of Incorporation", "date_of_incorporation"),
    ("State", "state"),
    ("Authorised Capital", "authorised_capital"),
    ("Paid Up Capital", "paid_up_capital"),
    ("Match Score", "match_score"),
    ("Match Method", "match_method"),
    ("Is Startup", "is_startup"),
    ("Registered Address", "registered_address"),
]


def generate_csv_bytes(companies: list[dict]) -> bytes:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[col[0] for col in EXPORT_COLUMNS], extrasaction="ignore")
    writer.writeheader()
    for company in companies:
        row = {col[0]: company.get(col[1], "") for col in EXPORT_COLUMNS}
        writer.writerow(row)
    return output.getvalue().encode("utf-8-sig")


def generate_xlsx_bytes(companies: list[dict], sheet_title: str = "Nexus Intel Export") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="F8FAFC", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="334155"),
        right=Side(style="thin", color="334155"),
        bottom=Side(style="thin", color="334155"),
    )

    headers = [col[0] for col in EXPORT_COLUMNS]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 30

    row_fill_even = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(vertical="center", wrap_text=False)

    for row_idx, company in enumerate(companies, 2):
        row_fill = row_fill_even if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for col_idx, (_, field) in enumerate(EXPORT_COLUMNS, 1):
            value = company.get(field, "")
            if isinstance(value, bool):
                value = "Yes" if value else "No"
            elif hasattr(value, "isoformat"):
                value = value.isoformat()
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    for col_idx, (header, _) in enumerate(EXPORT_COLUMNS, 1):
        max_len = len(header) + 2
        for row_idx in range(2, min(len(companies) + 2, 102)):
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, min(len(cell_val), 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len

    ws.freeze_panes = "A2"

    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Export Summary"])
    ws_summary.append(["Generated At", datetime.utcnow().isoformat()])
    ws_summary.append(["Total Records", len(companies)])
    ws_summary.append(["Matched Records", sum(1 for c in companies if c.get("match_score", 0) > 0)])
    ws_summary.append(["High Confidence (>90%)", sum(1 for c in companies if c.get("match_score", 0) >= 0.9)])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


async def create_and_upload_export(
    companies: list[dict],
    file_type: str,
    user_id: str,
    filter_params: Optional[dict] = None,
) -> dict:
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    file_name = f"nexus_intel_export_{timestamp}.{file_type}"
    r2_key = f"exports/{user_id}/{file_name}"

    tmp_path = os.path.join(tempfile.gettempdir(), file_name)
    try:
        if file_type == "csv":
            data = generate_csv_bytes(companies)
            content_type = "text/csv"
        else:
            data = generate_xlsx_bytes(companies)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        with open(tmp_path, "wb") as f:
            f.write(data)

        await upload_file_to_r2(tmp_path, r2_key, content_type)
        presigned_url = await generate_presigned_url(r2_key, expiry_seconds=86400)

        return {
            "file_name": file_name,
            "r2_key": r2_key,
            "r2_url": presigned_url,
            "file_size_bytes": len(data),
            "record_count": len(companies),
            "filter_params": filter_params or {},
        }
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
